"""
데이터베이스 스키마 추출을 위한 공통 유틸리티
"""
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from datetime import datetime

def apply_sheet_style(worksheet, df, startrow=0):
    """워크시트에 스타일을 적용합니다."""
    # 헤더 스타일 정의
    header_fill = PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid')
    header_font = Font(bold=True, color='000000')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    if not df.empty:
        # 헤더 추가
        for idx, col in enumerate(df.columns, 1):
            cell = worksheet.cell(row=startrow + 1, column=idx)
            cell.value = col
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        # 데이터 추가
        for r_idx, row in enumerate(df.values, startrow + 2):
            for c_idx, value in enumerate(row, 1):
                cell = worksheet.cell(row=r_idx, column=c_idx)
                cell.value = value
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = thin_border
                
                # 줄무늬 스타일
                if r_idx % 2 == 0:
                    cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        
        # 컬럼 너비 자동 조정
        for column in worksheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # 행 높이 설정
        worksheet.row_dimensions[startrow + 1].height = 25  # 헤더 행 높이
        for row in range(startrow + 2, worksheet.max_row + 1):
            worksheet.row_dimensions[row].height = 20  # 데이터 행 높이

def get_output_file_name(prefix):
    """타임스탬프가 포함된 출력 파일명을 생성합니다."""
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    return f"{prefix}_{timestamp}.xlsx"
