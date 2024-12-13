import pandas as pd
from sqlalchemy import create_engine, text
from datetime import datetime
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.styles.colors import Color
from openpyxl.worksheet.hyperlink import Hyperlink

def get_column_info(connection, table_name):
    """Get column information for a specific table using direct SQL queries"""
    column_query = text("""
        SELECT 
            c.COLUMN_NAME,
            c.DATA_TYPE,
            c.IS_NULLABLE,
            c.CHARACTER_MAXIMUM_LENGTH,
            c.NUMERIC_PRECISION,
            c.NUMERIC_SCALE,
            CASE WHEN pk.COLUMN_NAME IS NOT NULL THEN 'Y' ELSE 'N' END as IS_PRIMARY_KEY,
            CAST(ep.value AS NVARCHAR(4000)) as COLUMN_DESCRIPTION
        FROM INFORMATION_SCHEMA.COLUMNS c
        LEFT JOIN (
            SELECT ku.TABLE_CATALOG,ku.TABLE_SCHEMA,ku.TABLE_NAME,ku.COLUMN_NAME
            FROM INFORMATION_SCHEMA.TABLE_CONSTRAINTS tc
            JOIN INFORMATION_SCHEMA.KEY_COLUMN_USAGE ku
                ON tc.CONSTRAINT_TYPE = 'PRIMARY KEY' 
                AND tc.CONSTRAINT_NAME = ku.CONSTRAINT_NAME
                AND tc.TABLE_SCHEMA = ku.TABLE_SCHEMA
                AND tc.TABLE_NAME = ku.TABLE_NAME
        ) pk 
        ON c.TABLE_NAME = pk.TABLE_NAME 
        AND c.COLUMN_NAME = pk.COLUMN_NAME
        AND c.TABLE_SCHEMA = pk.TABLE_SCHEMA
        LEFT JOIN sys.columns sc
            ON sc.name = c.COLUMN_NAME
            AND sc.object_id = OBJECT_ID(c.TABLE_SCHEMA + '.' + c.TABLE_NAME)
        LEFT JOIN sys.extended_properties ep
            ON ep.major_id = sc.object_id
            AND ep.minor_id = sc.column_id
            AND ep.name = 'MS_Description'
        WHERE c.TABLE_SCHEMA = 'dbo'
        AND c.TABLE_NAME = :table_name
        ORDER BY c.ORDINAL_POSITION
    """)
    
    fk_query = text("""
        SELECT 
            COL_NAME(fk_cols.parent_object_id, fk_cols.parent_column_id) as parent_column,
            CONCAT(
                OBJECT_SCHEMA_NAME(fk.referenced_object_id), '.',
                OBJECT_NAME(fk.referenced_object_id), '.',
                COL_NAME(fk_cols.referenced_object_id, fk_cols.referenced_column_id)
            ) as referenced_table_column
        FROM sys.foreign_keys fk
        INNER JOIN sys.foreign_key_columns fk_cols ON 
            fk.object_id = fk_cols.constraint_object_id
        WHERE OBJECT_NAME(fk.parent_object_id) = :table_name
    """)
    
    columns = []
    
    # Execute queries
    results = connection.execute(column_query, {"table_name": table_name}).fetchall()
    fk_results = connection.execute(fk_query, {"table_name": table_name}).fetchall()
    
    # Convert foreign key results to dictionary for easier lookup
    fk_dict = {row[0]: row[1] for row in fk_results}
    
    for row in results:
        data_type = row.DATA_TYPE
        if row.CHARACTER_MAXIMUM_LENGTH:
            data_type += f'({row.CHARACTER_MAXIMUM_LENGTH})'
        elif row.NUMERIC_PRECISION is not None and row.NUMERIC_SCALE is not None:
            data_type += f'({row.NUMERIC_PRECISION},{row.NUMERIC_SCALE})'
            
        column_info = {
            '테이블명': table_name,
            '컬럼명': row.COLUMN_NAME,
            '데이터 타입': data_type,
            'Nullable': 'Y' if row.IS_NULLABLE == 'YES' else 'N',
            'PK': row.IS_PRIMARY_KEY,
            'FK': 'Y' if row.COLUMN_NAME in fk_dict else 'N',
            'FK 참조': fk_dict.get(row.COLUMN_NAME, ''),
            '설명': str(row.COLUMN_DESCRIPTION) if row.COLUMN_DESCRIPTION else ''
        }
        columns.append(column_info)
    
    return columns

def get_table_indexes(connection, table_name):
    """테이블의 인덱스 정보를 가져옵니다."""
    index_query = text("""
        SELECT 
            i.name AS index_name,
            STUFF((
                SELECT ', ' + c.name
                FROM sys.index_columns ic
                JOIN sys.columns c ON ic.object_id = c.object_id AND ic.column_id = c.column_id
                WHERE ic.object_id = i.object_id AND ic.index_id = i.index_id
                ORDER BY ic.key_ordinal
                FOR XML PATH('')
            ), 1, 2, '') AS columns,
            i.is_unique,
            i.is_primary_key,
            i.type_desc
        FROM sys.indexes i
        WHERE i.object_id = OBJECT_ID(:table_name)
        AND i.name IS NOT NULL
        ORDER BY i.name
    """)
    
    return connection.execute(index_query, {"table_name": table_name}).fetchall()

def get_views(connection):
    """모든 뷰 정보를 가져옵니다."""
    view_query = text("""
        SELECT 
            v.name AS view_name,
            OBJECT_DEFINITION(v.object_id) AS view_definition
        FROM sys.views v
        WHERE schema_name(v.schema_id) = 'dbo'
        ORDER BY v.name
    """)
    
    return connection.execute(view_query).fetchall()

def create_table_specification(connection_string, output_file):
    """
    데이터베이스 스키마를 읽어서 Excel 형식의 테이블 명세서를 생성합니다.
    
    Args:
        connection_string (str): 데이터베이스 연결 문자열
        output_file (str): 출력할 Excel 파일 경로
    """
    engine = create_engine(connection_string)
    
    with engine.connect() as connection:
        # 테이블 목록 가져오기
        table_query = text("""
            SELECT TABLE_NAME 
            FROM INFORMATION_SCHEMA.TABLES 
            WHERE TABLE_SCHEMA = 'dbo' 
            AND TABLE_TYPE = 'BASE TABLE' 
            ORDER BY TABLE_NAME
        """)
        
        # 모든 데이터 먼저 가져오기
        tables = connection.execute(table_query).fetchall()
        views = get_views(connection)
        
        # Excel 파일 생성
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # 1. 목차 시트 생성
            toc_data = {
                '구분': ['테이블'] * len(tables) + ['뷰'] * len(views),
                '이름': [t[0] for t in tables] + [v[0] for v in views],
                '설명': [''] * (len(tables) + len(views))
            }
            toc_df = pd.DataFrame(toc_data)
            toc_df.to_excel(writer, sheet_name='목차', index=False)
            
            # 목차 시트 스타일 적용
            ws_toc = writer.sheets['목차']
            apply_sheet_style(ws_toc, toc_df)
            
            # 2. 각 테이블 시트 생성
            for table in tables:
                table_name = table[0]
                
                # 2.1 컬럼 정보
                columns_info = get_column_info(connection, table_name)
                columns_df = pd.DataFrame(columns_info)
                
                # 컬럼 순서 재정의
                columns_df = columns_df[[
                    '테이블명',
                    '컬럼명',
                    '데이터 타입',
                    'Nullable',
                    'PK',
                    'FK',
                    'FK 참조',
                    '설명'  # 코멘트 컬럼 추가
                ]]
                
                # 2.2 인덱스 정보
                indexes = get_table_indexes(connection, table_name)
                indexes_df = pd.DataFrame([{
                    '인덱스명': idx.index_name,
                    '컬럼': idx.columns,
                    'Unique': 'Y' if idx.is_unique else 'N',
                    'PK': 'Y' if idx.is_primary_key else 'N',
                    '타입': idx.type_desc
                } for idx in indexes])
                
                # 시트 생성
                columns_df.to_excel(writer, sheet_name=table_name, index=False, startrow=2)
                
                # 시트 스타일 적용
                ws = writer.sheets[table_name]
                
                # 목차로 돌아가기 링크 추가
                ws['A1'] = '목차로 돌아가기'
                ws['A1'].hyperlink = '#목차!A1'
                ws['A1'].font = Font(color="0563C1", underline="single")
                ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
                
                # 스타일 적용
                apply_sheet_style(ws, columns_df, startrow=2)
                
                # 인덱스 정보 추가
                if not indexes_df.empty:
                    indexes_df.to_excel(writer, sheet_name=table_name, index=False, startrow=len(columns_df) + 5)
                    apply_sheet_style(ws, indexes_df, startrow=len(columns_df) + 5)
            
            # 3. 뷰 시트 생성
            if views:
                views_df = pd.DataFrame([{
                    '뷰명': v.view_name,
                    '정의': v.view_definition
                } for v in views])
                views_df.to_excel(writer, sheet_name='Views', index=False)
                
                # 뷰 시트 스타일 적용
                ws_views = writer.sheets['Views']
                apply_sheet_style(ws_views, views_df)
            
            # 4. 목차에 하이퍼링크 추가
            ws_toc = writer.sheets['목차']
            for idx, name in enumerate(toc_df['이름'], 2):  # 2부터 시작 (헤더 다음 행부터)
                cell = ws_toc.cell(row=idx, column=2)  # '이름' 컬럼
                cell.hyperlink = f"#{name}!A1"
                cell.font = Font(color="0563C1", underline="single")

def apply_sheet_style(worksheet, df, startrow=0):
    """워크시트에 스타일을 적용합니다."""
    # 테두리 스타일
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 헤더 스타일
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    
    # 데이터 행 스타일
    alternate_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    
    # 모든 셀에 테두리와 정렬 적용
    for row in worksheet.iter_rows(min_row=startrow+1, 
                                 max_row=startrow+len(df)+1, 
                                 min_col=1, 
                                 max_col=len(df.columns)):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # 헤더 스타일 적용
    for cell in list(worksheet.rows)[startrow:startrow+1][0]:
        cell.fill = header_fill
        cell.font = header_font
    
    # 줄무늬 패턴 적용 (홀수 행)
    for row_idx in range(startrow+2, startrow+len(df)+2, 2):
        for cell in list(worksheet.rows)[row_idx-1]:
            cell.fill = alternate_fill
    
    # 컬럼 너비 자동 조정
    for idx, column in enumerate(worksheet.columns):
        max_length = 0
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # 최대 너비 50으로 제한
        worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
    
    # 행 높이 설정
    worksheet.row_dimensions[startrow+1].height = 25  # 헤더 행 높이
    for row in range(startrow+2, startrow+len(df)+2):
        worksheet.row_dimensions[row].height = 20  # 데이터 행 높이

if __name__ == "__main__":
    connection_string = "mssql+pyodbc://smart_mes:smart_mes@127.0.0.1:1433/smart_mes?driver=SQL+Server"
    
    # 파일명에 타임스탬프 추가
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_file = f"table_specification_{timestamp}.xlsx"
    
    try:
        create_table_specification(connection_string, output_file)
    except PermissionError as e:
        print(f"Error: Excel 파일 '{output_file}'에 접근할 수 없습니다. ({str(e)})")
        print("파일이 이미 열려있다면 닫아주시기 바랍니다.")
    except Exception as e:
        print(f"Error: {str(e)}")
