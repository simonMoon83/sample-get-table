"""
오라클 데이터베이스 스키마를 Excel 명세서로 추출
"""
import pandas as pd
from sqlalchemy import create_engine, text
from db_schema_utils import apply_sheet_style, get_output_file_name
from openpyxl.styles import Font, Alignment

def get_column_info(connection, table_name, owner):
    """Get column information for a specific table using direct SQL queries"""
    column_query = text("""
        SELECT 
            c.COLUMN_NAME,
            c.DATA_TYPE,
            CASE 
                WHEN c.DATA_TYPE = 'NUMBER' AND c.DATA_PRECISION IS NOT NULL 
                THEN c.DATA_PRECISION || ',' || c.DATA_SCALE
                WHEN c.DATA_TYPE LIKE '%CHAR%' OR c.DATA_TYPE = 'NVARCHAR2'
                THEN c.CHAR_LENGTH
                ELSE NULL 
            END as DATA_LENGTH,
            c.NULLABLE,
            CASE WHEN p.COLUMN_NAME IS NOT NULL THEN 'Y' ELSE 'N' END as IS_PRIMARY_KEY,
            cc.COMMENTS
        FROM ALL_TAB_COLUMNS c
        LEFT JOIN (
            SELECT cols.TABLE_NAME, cols.COLUMN_NAME
            FROM ALL_CONSTRAINTS cons
            JOIN ALL_CONS_COLUMNS cols ON cons.CONSTRAINT_NAME = cols.CONSTRAINT_NAME
            WHERE cons.CONSTRAINT_TYPE = 'P'
            AND cons.OWNER = :owner
        ) p ON c.TABLE_NAME = p.TABLE_NAME AND c.COLUMN_NAME = p.COLUMN_NAME
        LEFT JOIN ALL_COL_COMMENTS cc 
            ON c.TABLE_NAME = cc.TABLE_NAME 
            AND c.COLUMN_NAME = cc.COLUMN_NAME
            AND cc.OWNER = :owner
        WHERE c.TABLE_NAME = :table_name
        AND c.OWNER = :owner
        ORDER BY c.COLUMN_ID
    """)
    
    fk_query = text("""
        SELECT 
            cols.COLUMN_NAME,
            cons.R_OWNER || '.' || cons.R_TABLE_NAME || '.' || rcols.COLUMN_NAME as REFERENCED_TABLE
        FROM ALL_CONSTRAINTS cons
        JOIN ALL_CONS_COLUMNS cols 
            ON cons.CONSTRAINT_NAME = cols.CONSTRAINT_NAME
            AND cons.OWNER = cols.OWNER
        JOIN ALL_CONS_COLUMNS rcols 
            ON cons.R_CONSTRAINT_NAME = rcols.CONSTRAINT_NAME
            AND cons.R_OWNER = rcols.OWNER
        WHERE cons.CONSTRAINT_TYPE = 'R'
        AND cons.OWNER = :owner
        AND cons.TABLE_NAME = :table_name
    """)
    
    # Execute queries
    results = connection.execute(column_query, {"table_name": table_name, "owner": owner}).fetchall()
    fk_results = connection.execute(fk_query, {"table_name": table_name, "owner": owner}).fetchall()
    
    # Convert foreign key results to dictionary for easier lookup
    fk_dict = {row[0]: row[1] for row in fk_results}
    
    columns = []
    for row in results:
        data_type = row.DATA_TYPE
        if row.DATA_LENGTH:
            data_type = f"{data_type}({row.DATA_LENGTH})"
            
        column_info = {
            '테이블명': table_name,
            '컬럼명': row.COLUMN_NAME,
            '데이터 타입': data_type,
            'Nullable': 'Y' if row.NULLABLE == 'Y' else 'N',
            'PK': row.IS_PRIMARY_KEY,
            'FK': 'Y' if row.COLUMN_NAME in fk_dict else 'N',
            'FK 참조': fk_dict.get(row.COLUMN_NAME, ''),
            '설명': row.COMMENTS if row.COMMENTS else ''
        }
        columns.append(column_info)
    
    return columns

def get_table_indexes(connection, table_name, owner):
    """테이블의 인덱스 정보를 가져옵니다."""
    index_query = text("""
        SELECT 
            i.INDEX_NAME,
            i.UNIQUENESS,
            LISTAGG(c.COLUMN_NAME, ', ') WITHIN GROUP (ORDER BY c.COLUMN_POSITION) as COLUMNS
        FROM ALL_INDEXES i
        JOIN ALL_IND_COLUMNS c ON i.INDEX_NAME = c.INDEX_NAME 
            AND i.TABLE_NAME = c.TABLE_NAME
            AND i.OWNER = c.INDEX_OWNER
        WHERE i.TABLE_NAME = :table_name
        AND i.OWNER = :owner
        GROUP BY i.INDEX_NAME, i.UNIQUENESS
        ORDER BY i.INDEX_NAME
    """)
    
    return connection.execute(index_query, {"table_name": table_name, "owner": owner}).fetchall()

def get_views(connection, owner):
    """모든 뷰 정보를 가져옵니다."""
    view_query = text("""
        SELECT 
            VIEW_NAME,
            TEXT as VIEW_DEFINITION,
            COMMENTS
        FROM ALL_VIEWS v
        LEFT JOIN ALL_TAB_COMMENTS c 
            ON v.VIEW_NAME = c.TABLE_NAME 
            AND v.OWNER = c.OWNER
        WHERE v.OWNER = :owner
        ORDER BY VIEW_NAME
    """)
    
    return connection.execute(view_query, {"owner": owner}).fetchall()

def create_table_specification(connection_string, owner):
    """
    데이터베이스 스키마를 읽어서 Excel 형식의 테이블 명세서를 생성합니다.
    
    Args:
        connection_string (str): 데이터베이스 연결 문자열
        owner (str): 스키마 소유자
    """
    engine = create_engine(connection_string)
    output_file = get_output_file_name("oracle_schema")
    
    with engine.connect() as connection:
        # 테이블 목록 가져오기
        table_query = text("""
            SELECT 
                t.TABLE_NAME,
                c.COMMENTS as TABLE_COMMENT
            FROM ALL_TABLES t
            LEFT JOIN ALL_TAB_COMMENTS c 
                ON t.TABLE_NAME = c.TABLE_NAME 
                AND t.OWNER = c.OWNER
            WHERE t.OWNER = :owner
            ORDER BY t.TABLE_NAME
        """)
        
        # 모든 데이터 먼저 가져오기
        tables = connection.execute(table_query, {"owner": owner}).fetchall()
        views = get_views(connection, owner)
        
        # Excel 파일 생성
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # 1. 목차 시트 생성
            toc_data = {
                '구분': ['테이블'] * len(tables) + ['뷰'] * len(views),
                '이름': [t[0] for t in tables] + [v[0] for v in views],
                '설명': [t[1] if t[1] else '' for t in tables] + [v[2] if v[2] else '' for v in views]
            }
            toc_df = pd.DataFrame(toc_data)
            toc_df.to_excel(writer, sheet_name='목차', index=False)
            
            # 목차 시트 스타일 적용
            ws_toc = writer.sheets['목차']
            apply_sheet_style(ws_toc, toc_df)
            
            # 2. 각 테이블 시트 생성
            for table in tables:
                table_name = table[0]
                
                # 2.1 컬럼 정보
                columns_info = get_column_info(connection, table_name, owner)
                columns_df = pd.DataFrame(columns_info)
                
                # 컬럼 순서 재정의
                columns_df = columns_df[[
                    '테이블명',
                    '컬럼명',
                    '데이터 타입',
                    'Nullable',
                    'PK',
                    'FK',
                    'FK 참조',
                    '설명'  # 코멘트 컬럼 추가
                ]]
                
                # 2.2 인덱스 정보
                indexes = get_table_indexes(connection, table_name, owner)
                indexes_df = pd.DataFrame([{
                    '인덱스명': idx.INDEX_NAME,
                    '컬럼': idx.COLUMNS,
                    'Unique': 'Y' if idx.UNIQUENESS == 'UNIQUE' else 'N'
                } for idx in indexes])
                
                # 시트 생성
                columns_df.to_excel(writer, sheet_name=table_name, index=False, startrow=2)
                
                # 시트 스타일 적용
                ws = writer.sheets[table_name]
                
                # 목차로 돌아가기 링크 추가
                ws['A1'] = '목차로 돌아가기'
                ws['A1'].hyperlink = '#목차!A1'
                ws['A1'].font = Font(color="0563C1", underline="single")
                ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
                
                # 스타일 적용
                apply_sheet_style(ws, columns_df, startrow=2)
                
                # 인덱스 정보 추가
                if not indexes_df.empty:
                    indexes_df.to_excel(writer, sheet_name=table_name, index=False, startrow=len(columns_df) + 5)
                    apply_sheet_style(ws, indexes_df, startrow=len(columns_df) + 5)
            
            # 3. 뷰 시트 생성
            if views:
                views_df = pd.DataFrame([{
                    '뷰명': v.VIEW_NAME,
                    '설명': v.COMMENTS if v.COMMENTS else '',
                    '정의': v.VIEW_DEFINITION
                } for v in views])
                views_df.to_excel(writer, sheet_name='Views', index=False)
                
                # 뷰 시트 스타일 적용
                ws_views = writer.sheets['Views']
                apply_sheet_style(ws_views, views_df)
    
    print(f"오라클 테이블 명세서가 생성되었습니다: {output_file}")

if __name__ == "__main__":
    # 오라클 연결 문자열 예시 (JDBC Thin style)
    connection_string = "oracle+oracledb://username:password@hostname:1521/service_name"
    owner = "SCHEMA_NAME"  # 스키마 이름을 대문자로 지정
    
    try:
        create_table_specification(connection_string, owner)
    except Exception as e:
        print(f"Error: {str(e)}")
